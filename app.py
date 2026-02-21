from flask import Flask, render_template, jsonify, request, redirect, url_for, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
from werkzeug.security import check_password_hash, generate_password_hash
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
import io
import openpyxl
from copy import copy
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from num2words import num2words
import os
import re
import easyocr
from werkzeug.utils import secure_filename
# --- App Initialization and Configuration ---
# --- App Initialization and Configuration ---
app = Flask(__name__)
app.config['SECRET_KEY'] = 'ENTER_SECRET_KEY' 

# --- RELIABLE DATABASE PATH FOR DESKTOP APP ---
# Gets the user's home directory (e.g., C:/Users/YourUser)
user_home = os.path.expanduser('~')
# Create a dedicated folder for your app's data
app_data_dir = os.path.join(user_home, '.BillGeneratorApp')
os.makedirs(app_data_dir, exist_ok=True)
db_path = os.path.join(app_data_dir, 'database.db')
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{db_path}'
# --- END OF CHANGE ---

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# --- NEW OCR CONFIGURATION ---
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'}

# Create upload folder if it doesn't exist
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
    
    
db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login' # Redirect to login page if user is not authenticated

# --- Admin User Configuration ---
# Set your desired admin username.
ADMIN_USERNAME = 'ENTER_USERNAME'
# This is a secure "hash" of the default password 'admin'.
# To change your password, open a Python terminal and run:
# from werkzeug.security import generate_password_hash
# print(generate_password_hash('admin'))
# from werkzeug.security import generate_password_hash
# print(generate_password_hash('YOUR_NEW_PASSWORD_HERE'))
# Then, copy the long string output and paste it here.
ADMIN_PASSWORD_HASH = 'scrypt:32768:8:1$X2JhJDEde3NRmfyv$c46c363dfe48b6b609a81cd188b2d89de9c7ca1489d38d720a9daac3d8c570f80d22b7f0d75761b363379f6f7ffd05f9013fb52572c1100435ed65f5b90d9e12' 

# --- Simplified User Class (In-Memory) ---
# This class represents the logged-in user without needing a database table.
class AdminUser(UserMixin):
    def __init__(self, id, username):
        self.id = id
        self.username = username

# --- 1. Define the Database Models (Tables) ---

# Model for individual, reusable items
class Item(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    item_name = db.Column(db.String(100), unique=True, nullable=False)
    unit = db.Column(db.String(50), nullable=False, default='Nos') # ADD THIS
    hsn = db.Column(db.String(50), nullable=True) # ADD THIS
    cgst_percentage = db.Column(db.Float, nullable=False)
    sgst_percentage = db.Column(db.Float, nullable=False)

# Model for the main Bill document
class Bill(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    bill_number = db.Column(db.String(50), unique=True, nullable=False)
    date = db.Column(db.DateTime, default=datetime.utcnow)
    grand_total = db.Column(db.Float, nullable=False)
    payment_status = db.Column(db.String(20), nullable=False, default='Pending')
    items = db.relationship('BillItem', backref='bill', cascade="all, delete-orphan")


# Model for each line item within a Bill
class BillItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    bill_id = db.Column(db.Integer, db.ForeignKey('bill.id'), nullable=False)
    item_name = db.Column(db.String(100), nullable=False)
    quantity = db.Column(db.Float, nullable=False)
    unit = db.Column(db.String(50), nullable=False, default='Nos') # ADD THIS
    hsn = db.Column(db.String(50), nullable=True) # ADD THIS
    price = db.Column(db.Float, nullable=False)
    cgst_percent = db.Column(db.Float, nullable=False)
    sgst_percent = db.Column(db.Float, nullable=False)
    total_amount = db.Column(db.Float, nullable=False)
    
# Add this new model
class Task(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    description = db.Column(db.String(200), nullable=False)
    due_date = db.Column(db.DateTime, nullable=False)
    completed = db.Column(db.Boolean, default=False)


# Add this new route for the Payment Tracker
@app.route('/payment_tracker')
@login_required
def payment_tracker():
    return render_template('payment-tracker.html')

# --- User Loader for Flask-Login ---
@login_manager.user_loader
def load_user(user_id):
    if int(user_id) == 1:
        return AdminUser(id=1, username=ADMIN_USERNAME)
    return None

# --- Authentication Routes ---

# Find and replace this entire function
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('home')) # Changed from 'bill_management'
        
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        if username == ADMIN_USERNAME and check_password_hash(ADMIN_PASSWORD_HASH, password):
            admin = AdminUser(id=1, username=ADMIN_USERNAME)
            login_user(admin)
            # This is the main change: redirect to 'home' after login
            return redirect(url_for('home'))
        else:
            flash('Invalid username or password.', 'error')
            
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# --- 2. HTML Page Routes ---

# Find and replace this entire function
@app.route('/')
@login_required
def home():
    # This now renders your index.html page instead of redirecting
    return render_template('index.html')

@app.route('/bill_management')
@login_required
def bill_management():
    return render_template('bill_management.html')

@app.route('/bill-details')
@login_required
def bill_details():
    bill_id = request.args.get('id')
    return render_template('bill-details.html', bill_id=bill_id)

@app.route('/new_bill')
@login_required
def new_bill():
    return render_template('new-bill.html')

# Add this new route
@app.route('/todo')
@login_required
def todo():
    return render_template('todo.html')

# Add this new route for the database page
@app.route('/database')
@login_required
def database():
    return render_template('database.html')

# --- 3. API Routes for Data Handling ---

# Add this new API endpoint for updating payment status
@app.route('/api/bills/<int:bill_id>/status', methods=['PUT'])
@login_required
def update_payment_status(bill_id):
    bill = Bill.query.get_or_404(bill_id)
    data = request.get_json()
    
    new_status = data.get('status')
    if new_status not in ['Pending', 'Paid']:
        return jsonify({'error': 'Invalid status'}), 400
        
    bill.payment_status = new_status
    db.session.commit()
    return jsonify({'message': f'Bill {bill.bill_number} status updated to {new_status}.'})

# Add these three new API endpoints for managing the Item table

# GET all items for the database page
@app.route('/api/items', methods=['GET'])
@login_required
def get_all_items():
    items = Item.query.order_by(Item.item_name).all()
    output = []
    for item in items:
        item_data = {
            'id': item.id,
            'item_name': item.item_name,
            'unit': item.unit, # ADD THIS
            'hsn': item.hsn,   # ADD THIS
            'cgst': item.cgst_percentage,
            'sgst': item.sgst_percentage
        }
        output.append(item_data)
    return jsonify(output)

# PUT (update) an existing item
@app.route('/api/items/<int:item_id>', methods=['PUT'])
@login_required
def update_item(item_id):
    item = Item.query.get_or_404(item_id)
    data = request.get_json()
    
    item.item_name = data['name']
    item.unit = data['unit']       # ADD THIS
    item.hsn = data.get('hsn')     # ADD THIS
    item.cgst_percentage = data['cgst']
    item.sgst_percentage = data['sgst']
    
    db.session.commit()
    return jsonify({'message': 'Item updated successfully!'})

# DELETE an item
@app.route('/api/items/<int:item_id>', methods=['DELETE'])
@login_required
def delete_item(item_id):
    item = Item.query.get_or_404(item_id)
    db.session.delete(item)
    db.session.commit()
    return jsonify({'message': 'Item deleted successfully!'})


# --- (The rest of your API routes continue below) ---

# --- 3. API Routes for Data Handling ---

@app.route('/api/search')
@login_required
def search():
    query = request.args.get('q')
    search_term = f"{query}%"
    items = Item.query.filter(Item.item_name.ilike(search_term)).all()
    suggestions = [item.item_name for item in items]
    return jsonify(suggestions)

@app.route('/api/item/<item_name>')
@login_required
def get_item_details(item_name):
    item = Item.query.filter_by(item_name=item_name).first()
    if item:
        return jsonify({
            'cgst': item.cgst_percentage, 
            'sgst': item.sgst_percentage,
            'unit': item.unit, # ADD THIS
            'hsn': item.hsn    # ADD THIS
        })
    else:
        return jsonify({'error': 'Item not found'}), 404

@app.route('/api/item', methods=['POST'])
@login_required
def add_item():
    data = request.get_json()
    existing_item = Item.query.filter_by(item_name=data['name']).first()
    if existing_item:
        return jsonify({'error': 'Item already exists'}), 409

    new_item = Item(
        item_name=data['name'],
        unit=data['unit'],       # ADD THIS
        hsn=data.get('hsn'),     # ADD THIS
        cgst_percentage=data['cgst'],
        sgst_percentage=data['sgst']
    )
    db.session.add(new_item)
    db.session.commit()
    return jsonify({'message': 'Item added successfully'}), 201

@app.route('/api/bills', methods=['GET'])
@login_required
def get_all_bills():
    bills = Bill.query.order_by(Bill.date.desc()).all()
    output = []
    for bill in bills:
        total_items = len(bill.items)
        bill_data = {
            'id': bill.id,
            'bill_number': bill.bill_number,
            # This check prevents a crash if a bill's date is missing
            'date': bill.date.strftime('%b %d, %Y') if bill.date else 'No Date',
            'total_items': total_items,
            # This check prevents a crash if grand_total is missing
            'total_amount': bill.grand_total if bill.grand_total is not None else 0,
            'payment_status': bill.payment_status
        }
        output.append(bill_data)
    return jsonify(output)

@app.route('/api/bills/<int:bill_id>', methods=['GET'])
@login_required
def get_single_bill(bill_id):
    bill = Bill.query.get_or_404(bill_id)
    bill_items = []
    for item in bill.items:
        bill_items.append({
            'item_name': item.item_name,
            'quantity': item.quantity,
            'unit': item.unit, # ADDED
            'hsn': item.hsn,   # ADDED
            'price': item.price,
            'cgst_percent': item.cgst_percent,
            'sgst_percent': item.sgst_percent,
            'total_amount': item.total_amount
        })
    
    return jsonify({
        'bill_number': bill.bill_number,
        'date': bill.date.strftime('%B %d, %Y, %I:%M %p'),
        'grand_total': bill.grand_total,
        'items': bill_items
    })

@app.route('/api/bills/next_bill_number')
@login_required
def get_next_bill_number():
    base_name = "NEW BILL"
    if not Bill.query.filter_by(bill_number=base_name).first():
        return jsonify({'next_bill_number': base_name})
    
    counter = 1
    while True:
        next_name = f"{base_name}({counter})"
        if not Bill.query.filter_by(bill_number=next_name).first():
            return jsonify({'next_bill_number': next_name})
        counter += 1
        
@app.route('/api/bills', methods=['POST'])
@login_required
def create_bill():
    data = request.get_json()
    
    new_bill = Bill(
        bill_number=data['bill_number'],
        grand_total=data['grand_total']
    )
    db.session.add(new_bill)
    
    for item_data in data['items']:
        bill_item = BillItem(
            bill=new_bill,
            item_name=item_data['name'],
            quantity=item_data['quantity'],
            unit=item_data['unit'],       # ADDED
            hsn=item_data.get('hsn'),     # ADDED
            price=item_data['price'],
            cgst_percent=item_data['cgst'],
            sgst_percent=item_data['sgst'],
            total_amount=item_data['final_total']
        )
        db.session.add(bill_item)
        
    db.session.commit()
    return jsonify({'message': 'Bill created successfully!', 'id': new_bill.id}), 201


@app.route('/api/bills/<int:bill_id>/add_items', methods=['POST'])
@login_required
def add_items_to_bill(bill_id):
    bill = Bill.query.get_or_404(bill_id)
    items_to_add = request.get_json()
    
    if not isinstance(items_to_add, list):
        return jsonify({'error': 'Request body must be a list of items'}), 400

    total_added_value = 0
    for item_data in items_to_add:
        # Recalculate amounts on the server for security and consistency
        taxable_value = float(item_data['quantity']) * float(item_data['price'])
        cgst_amount = taxable_value * (float(item_data['cgst']) / 100)
        sgst_amount = taxable_value * (float(item_data['sgst']) / 100)
        final_total = taxable_value + cgst_amount + sgst_amount
        total_added_value += final_total
        
        new_item = BillItem(
            bill_id=bill.id,
            item_name=item_data['name'],
            quantity=item_data['quantity'],
            unit=item_data['unit'],       # ADD THIS LINE
            hsn=item_data.get('hsn'),     # ADD THIS LINE
            price=item_data['price'],
            cgst_percent=item_data['cgst'],
            sgst_percent=item_data['sgst'],
            total_amount=final_total
        )
        db.session.add(new_item)

    # Update the grand total of the target bill
    bill.grand_total += total_added_value
    db.session.commit()
    
    return jsonify({'message': f'{len(items_to_add)} items added to bill {bill.bill_number} successfully!'})


@app.route('/api/bills/<int:bill_id>', methods=['PUT'])
@login_required
def update_bill(bill_id):
    bill = Bill.query.get_or_404(bill_id)
    data = request.get_json()

    bill.bill_number = data['bill_number']
    bill.grand_total = data['grand_total']
    bill.date = datetime.utcnow()

    BillItem.query.filter_by(bill_id=bill_id).delete()
    
    for item_data in data['items']:
        bill_item = BillItem(
            bill_id=bill_id,
            item_name=item_data['name'],
            quantity=item_data['quantity'],
            unit=item_data['unit'],   # ADDED
            hsn=item_data.get('hsn'), # ADDED
            price=item_data['price'],
            cgst_percent=item_data['cgst'],
            sgst_percent=item_data['sgst'],
            total_amount=item_data['final_total']
        )
        db.session.add(bill_item)
        
    db.session.commit()
    return jsonify({'message': f'Bill {bill_id} updated successfully!'})

@app.route('/api/bills/<int:bill_id>', methods=['DELETE'])
@login_required
def delete_bill(bill_id):
    bill = Bill.query.get_or_404(bill_id)
    db.session.delete(bill)
    db.session.commit()
    return jsonify({'message': 'Bill deleted successfully!'})

# Add this new route to your app.py file

@app.route('/api/master-item/update', methods=['PUT'])
@login_required
def update_master_item():
    data = request.get_json()
    item_name = data.get('name')

    if not item_name:
        return jsonify({'error': 'Item name is required'}), 400

    # Find the item in the master Item table
    item = Item.query.filter_by(item_name=item_name).first()

    if not item:
        return jsonify({'error': 'Master item not found'}), 404

    # Update only the specified fields
    if 'hsn' in data:
        item.hsn = data['hsn']
    if 'cgst' in data:
        item.cgst_percentage = data['cgst']
    if 'sgst' in data:
        item.sgst_percentage = data['sgst']
    
    db.session.commit()
    return jsonify({'message': f"'{item_name}' updated in the master database."})


# --- API Routes for Tasks ---

@app.route('/api/tasks', methods=['GET'])
@login_required
def get_tasks():
    tasks = Task.query.order_by(Task.due_date).all()
    output = []
    for task in tasks:
        task_data = {
            'id': task.id,
            'description': task.description,
            'due_date': task.due_date.isoformat(),
            'completed': task.completed
        }
        output.append(task_data)
    return jsonify(output)

@app.route('/api/tasks', methods=['POST'])
@login_required
def add_task():
    data = request.get_json()
    new_task = Task(
        description=data['description'],
        due_date=datetime.fromisoformat(data['due_date'])
    )
    db.session.add(new_task)
    db.session.commit()
    return jsonify({'message': 'Task added successfully!', 'id': new_task.id}), 201

@app.route('/api/tasks/<int:task_id>', methods=['PUT'])
@login_required
def update_task(task_id):
    task = Task.query.get_or_404(task_id)
    data = request.get_json()
    
    # Update only the fields provided in the request
    if 'description' in data:
        task.description = data['description']
    if 'due_date' in data:
        task.due_date = datetime.fromisoformat(data['due_date'])
    if 'completed' in data:
        task.completed = data['completed']
        
    db.session.commit()
    return jsonify({'message': 'Task updated successfully!'})

@app.route('/api/tasks/<int:task_id>', methods=['DELETE'])
@login_required
def delete_task(task_id):
    task = Task.query.get_or_404(task_id)
    db.session.delete(task)
    db.session.commit()
    return jsonify({'message': 'Task deleted successfully!'})


@app.route('/api/bill/<int:bill_id>/download')
@login_required
def download_bill(bill_id):
    bill = Bill.query.get_or_404(bill_id)

    try:
        workbook = openpyxl.load_workbook('bill_template.xlsx')
        sheet = workbook.active
    except FileNotFoundError:
        return "Error: 'template.xlsx' not found. Make sure it is in the root project directory.", 404

    # --- THIS IS THE NEW, REWRITTEN LOGIC ---

    # 1. Define key rows and a standard border style
    start_row = 12
    template_row_for_style = 12
    thin_border_side = Side(style='thin')
    outside_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    # 2. Store original merged cells
    original_merged_cells = list(sheet.merged_cells.ranges)
    
    # 3. Determine how many rows to insert
    num_items = len(bill.items)
    rows_to_insert = num_items - 1 if num_items > 0 else 0

    # 4. If we need to insert rows, do all insertions at once
    if rows_to_insert > 0:
        sheet.insert_rows(start_row + 1, amount=rows_to_insert)

    # 5. Copy styles for all new rows
    for i in range(rows_to_insert):
        new_row_index = start_row + 1 + i
        for col_idx, cell in enumerate(sheet[template_row_for_style], 1):
            new_cell = sheet.cell(row=new_row_index, column=col_idx)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    # 6. Unmerge all original ranges, then re-merge them in their new shifted positions
    for merged_range in original_merged_cells:
        sheet.unmerge_cells(str(merged_range))
    
    for merged_range in original_merged_cells:
        min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
        if min_row > start_row:
            min_row += rows_to_insert
            max_row += rows_to_insert
        new_range_coord = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
        sheet.merge_cells(new_range_coord)

    # 7. Populate Header Information
    sheet['M6'] = bill.bill_number
    sheet['M9'] = bill.date.strftime('%d.%m.%Y')

    # 8. Write item data into the prepared rows
    total_taxable_value = 0
    total_cgst_amount = 0
    total_sgst_amount = 0

    for index, item in enumerate(bill.items):
        current_row = start_row + index
        
        # Calculations
        taxable_value = item.quantity * item.price
        cgst_amount = taxable_value * (item.cgst_percent / 100)
        sgst_amount = taxable_value * (item.sgst_percent / 100)
        
        total_taxable_value += taxable_value
        total_cgst_amount += cgst_amount
        total_sgst_amount += sgst_amount

        # Data to write in this row
        row_data = [
            index + 1, item.item_name, item.quantity, item.unit, item.hsn, item.price,
            taxable_value, taxable_value, item.cgst_percent, cgst_amount,
            item.sgst_percent, sgst_amount, item.total_amount
        ]

        # Write data and apply styles to each cell in the row
        for col_idx, value in enumerate(row_data, 1):
            cell = sheet.cell(row=current_row, column=col_idx, value=value)
            template_cell = sheet.cell(row=template_row_for_style, column=col_idx)
            # Ensure alignment and borders are explicitly set
            cell.alignment = copy(template_cell.alignment)
            cell.border = outside_border

    # 9. Write summary totals and apply styles
    total_row = start_row + num_items
    summary_data = {
        6: "Total", 8: total_taxable_value, 10: total_cgst_amount,
        12: total_sgst_amount, 13: bill.grand_total
    }
    for col_idx, value in summary_data.items():
        cell = sheet.cell(row=total_row, column=col_idx, value=value)
        template_cell = sheet.cell(row=template_row_for_style, column=col_idx)
        cell.alignment = copy(template_cell.alignment)
        cell.border = outside_border

    # 10. Populate the shifted summary section
    summary_start_row = 15 + rows_to_insert
    sheet[f'M{summary_start_row}'] = bill.grand_total
    sheet[f'M{summary_start_row + 2}'] = total_taxable_value
    sheet[f'M{summary_start_row + 3}'] = total_cgst_amount
    sheet[f'M{summary_start_row + 4}'] = total_sgst_amount
    sheet[f'M{summary_start_row + 6}'] = bill.grand_total

    # 11. Convert grand total to words
    words_row = 22 + rows_to_insert
    grand_total_in_words = f"{num2words(bill.grand_total, lang='en_IN').title()} Only"
    sheet[f'K{words_row}'] = grand_total_in_words

    # 12. Save and send the file
    file_stream = io.BytesIO()
    workbook.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name=f'{bill.bill_number}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


def allowed_file(filename):
    """Checks if the uploaded file has an allowed extension."""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def recognize_with_easyocr(image_path):
    """
    Recognizes text from an image using the easyocr library.
    """
    try:
        # Initialize the EasyOCR reader for English
        reader = easyocr.Reader(['en'])
        # Read the text from the image
        results = reader.readtext(image_path)
        # Extract just the text and combine it
        recognized_text = "\n".join([result[1] for result in results])
        return recognized_text
    except Exception as e:
        print(f"An error occurred during OCR: {e}")
        return None

def parse_ocr_text(text):
    """
    Parses raw OCR text to extract potential item names.
    This simple parser looks for lines of text that likely represent item descriptions.
    """
    items = []
    lines = text.split('\n')
    for line in lines:
        # This regex looks for lines that contain at least one letter and are 4 or more characters long.
        # It's a basic filter to remove noise like isolated numbers or symbols.
        if re.search(r'[a-zA-Z]', line) and len(line.strip()) >= 4:
            # We are primarily interested in the item name.
            # More complex logic could be added here to find quantity and price.
            item_name = line.strip()
            items.append({'name': item_name})
    return items

@app.route('/api/ocr/upload', methods=['POST'])
@login_required
def ocr_upload():
    """
    API endpoint to handle image upload, OCR processing, and cleanup.
    """
    if 'file' not in request.files:
        return jsonify({'error': 'No file part in the request'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
        
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        try:
            file.save(filepath)
            
            # Process the image with OCR
            recognized_text = recognize_with_easyocr(filepath)
            
            if recognized_text is None:
                return jsonify({'error': 'Could not process image with OCR.'}), 500

            # Parse the text to find items
            extracted_items = parse_ocr_text(recognized_text)
            
            return jsonify(extracted_items)
            
        except Exception as e:
            return jsonify({'error': f'An server error occurred: {e}'}), 500
        finally:
            # Clean up the uploaded file
            if os.path.exists(filepath):
                os.remove(filepath)
    
    return jsonify({'error': 'File type not allowed'}), 400
# --- END NEW OCR FUNCTIONS AND ROUTE ---

